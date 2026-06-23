from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Dict
from datetime import datetime
from io import BytesIO


class ExportService:
    """Service for exporting player data to Excel."""

    @staticmethod
    def detect_duplicates(players: List[Dict]) -> Dict[str, List[Dict]]:
        """
        Detect duplicate full names within the same team.
        Returns dict mapping team_id to list of duplicate player groups.
        """
        duplicates_by_team = {}

        for team_id, team_players in {}.items():
            # Group players by full name
            name_groups = {}
            for player in team_players:
                full_name = player.get("fullName") or f"{player.get('firstName', '')} {player.get('lastName', '')}"
                if full_name.strip():
                    if full_name not in name_groups:
                        name_groups[full_name] = []
                    name_groups[full_name].append(player)

            # Keep only duplicates
            duplicates = {name: group for name, group in name_groups.items() if len(group) > 1}
            if duplicates:
                duplicates_by_team[team_id] = duplicates

        return duplicates_by_team

    @staticmethod
    def create_xlsx(
        teams_data: List[Dict],
        filename_prefix: str = "verified_players"
    ) -> Tuple[BytesIO, str]:
        """
        Create Excel file with verified players, one sheet per team.
        Returns (BytesIO object, filename).
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for team_data in teams_data:
            team_name = team_data["name"]
            team_age_group = team_data["ageGroup"]
            team_gender = team_data["gender"]
            players = team_data["players"]

            if not players:
                continue

            # Create sheet with team name
            sheet = wb.create_sheet(title=team_name[:31])  # Excel sheet name max 31 chars

            # Set up headers
            headers = [
                "No",
                "Team",
                "Age Group",
                "Gender",
                "First Name",
                "Last Name",
                "Full Name",
                "Source File",
                "Verified Date",
            ]

            sheet.append(headers)

            # Style header row
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Add data rows
            for idx, player in enumerate(players, 1):
                row_data = [
                    idx,
                    team_name,
                    team_age_group,
                    team_gender,
                    player.get("firstName") or "",
                    player.get("lastName") or "",
                    player.get("fullName") or "",
                    player.get("sourceFilename") or "",
                    player.get("verifiedAt", "")[:10] if player.get("verifiedAt") else "",
                ]
                sheet.append(row_data)

            # Set column widths
            column_widths = {
                "A": 5,   # No
                "B": 15,  # Team
                "C": 12,  # Age Group
                "D": 12,  # Gender
                "E": 15,  # First Name
                "F": 15,  # Last Name
                "G": 25,  # Full Name
                "H": 20,  # Source File
                "I": 15,  # Verified Date
            }

            for col_letter, width in column_widths.items():
                sheet.column_dimensions[col_letter].width = width

            # Apply alternating row colors for readability
            light_fill = PatternFill(
                start_color="D9E8F5",
                end_color="D9E8F5",
                fill_type="solid"
            )
            regular_fill = PatternFill(fill_type=None)
            center_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for row_num in range(2, len(players) + 2):
                for col_num in range(1, len(headers) + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = center_alignment

                    # Alternate row colors
                    if row_num % 2 == 0:
                        cell.fill = light_fill
                    else:
                        cell.fill = regular_fill

            # Freeze header row
            sheet.freeze_panes = "A2"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"

        return output, filename

    @staticmethod
    def format_export_data(teams: List, db_players: List) -> List[Dict]:
        """
        Format team and player data for export.
        Groups players by team.
        """
        team_dict = {team.id: team for team in teams}
        export_data = []

        # Group players by team
        players_by_team = {}
        for player in db_players:
            if player.teamId not in players_by_team:
                players_by_team[player.teamId] = []
            players_by_team[player.teamId].append(player)

        # Create export structure
        for team in teams:
            if team.id in players_by_team:
                team_export = {
                    "id": team.id,
                    "name": team.name,
                    "ageGroup": team.ageGroup,
                    "gender": team.gender,
                    "players": [
                        {
                            "firstName": p.firstName,
                            "lastName": p.lastName,
                            "fullName": p.fullName,
                            "sourceFilename": p.sourceFilename,
                            "verifiedAt": p.verifiedAt.isoformat() if p.verifiedAt else None,
                        }
                        for p in players_by_team[team.id]
                    ]
                }
                export_data.append(team_export)

        return export_data
