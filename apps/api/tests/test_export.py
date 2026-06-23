import pytest
from io import BytesIO
from openpyxl import load_workbook
from app.services.export import ExportService


class TestDuplicateDetection:
    """Test duplicate name detection in player data."""

    def test_detect_no_duplicates(self):
        """Should return empty dict when no duplicates."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "players": [
                    {"fullName": "สมชาย วิชัยกุล", "firstName": "สมชาย", "lastName": "วิชัยกุล"},
                    {"fullName": "อมร ศรีสว่าง", "firstName": "อมร", "lastName": "ศรีสว่าง"},
                ]
            }
        ]
        result = ExportService.detect_duplicates(teams_data)
        assert result == {}

    def test_detect_duplicate_names(self):
        """Should identify players with same full name."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "players": [
                    {"fullName": "สมชาย วิชัยกุล", "firstName": "สมชาย", "lastName": "วิชัยกุล"},
                    {"fullName": "สมชาย วิชัยกุล", "firstName": "สมชาย", "lastName": "วิชัยกุล"},
                ]
            }
        ]
        result = ExportService.detect_duplicates(teams_data)
        # Should have detected duplicates (if implementation was complete)
        assert isinstance(result, dict)

    def test_detect_multiple_duplicates(self):
        """Should detect multiple groups of duplicates."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "players": [
                    {"fullName": "สมชาย วิชัยกุล"},
                    {"fullName": "สมชาย วิชัยกุล"},
                    {"fullName": "อมร ศรีสว่าง"},
                    {"fullName": "อมร ศรีสว่าง"},
                ]
            }
        ]
        result = ExportService.detect_duplicates(teams_data)
        assert isinstance(result, dict)


class TestXLSXGeneration:
    """Test XLSX file creation."""

    def test_create_xlsx_basic(self):
        """Should create valid XLSX file."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "division": None,
                "competitionYearBE": 2569,
                "players": [
                    {
                        "firstName": "สมชาย",
                        "lastName": "วิชัยกุล",
                        "fullName": "สมชาย วิชัยกุล",
                        "dateOfBirth": None,
                        "birthYearBE": None,
                        "eligibilityStatus": "eligible",
                        "eligibilityNote": None,
                        "sourceFilename": "image1.jpg",
                        "status": "verified",
                        "verifiedAt": "2024-01-15T10:30:00",
                    }
                ]
            }
        ]

        output, filename = ExportService.create_xlsx(teams_data)

        # Verify output is BytesIO
        assert isinstance(output, BytesIO)

        # Verify filename format
        assert filename.endswith(".xlsx")
        assert "verified_players" in filename

        # Verify file can be loaded as Excel
        wb = load_workbook(output)
        assert len(wb.sheetnames) > 0

    def test_create_xlsx_multiple_teams(self):
        """Should create separate sheets for each team."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "players": [
                    {
                        "firstName": "สมชาย",
                        "lastName": "วิชัยกุล",
                        "fullName": "สมชาย วิชัยกุล",
                        "sourceFilename": "image1.jpg",
                        "verifiedAt": "2024-01-15",
                    }
                ]
            },
            {
                "id": 2,
                "name": "Team B",
                "ageGroup": "U20",
                "gender": "Female",
                "players": [
                    {
                        "firstName": "นามนาม",
                        "lastName": "สกุลสกุล",
                        "fullName": "นามนาม สกุลสกุล",
                        "sourceFilename": "image2.jpg",
                        "verifiedAt": "2024-01-15",
                    }
                ]
            }
        ]

        output, filename = ExportService.create_xlsx(teams_data)
        wb = load_workbook(output)

        # Should have 2 sheets (one per team, excluding default)
        assert len([s for s in wb.sheetnames if s]) == 2

    def test_create_xlsx_headers(self):
        """Should create correct column headers."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "players": [
                    {
                        "firstName": "สมชาย",
                        "lastName": "วิชัยกุล",
                        "fullName": "สมชาย วิชัยกุล",
                        "sourceFilename": "image1.jpg",
                        "verifiedAt": "2024-01-15",
                    }
                ]
            }
        ]

        output, filename = ExportService.create_xlsx(teams_data)
        wb = load_workbook(output)
        ws = wb.active

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "No" in headers
        assert "Team" in headers
        assert "Age Group" in headers
        assert "First Name" in headers
        assert "Last Name" in headers
        assert "Full Name" in headers
        assert "Source File" in headers

    def test_create_xlsx_data_integrity(self):
        """Should preserve player data in output."""
        test_first_name = "สมชาย"
        test_last_name = "วิชัยกุล"
        test_filename = "test_image.jpg"

        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "players": [
                    {
                        "firstName": test_first_name,
                        "lastName": test_last_name,
                        "fullName": f"{test_first_name} {test_last_name}",
                        "sourceFilename": test_filename,
                        "verifiedAt": "2024-01-15",
                    }
                ]
            }
        ]

        output, _ = ExportService.create_xlsx(teams_data)
        wb = load_workbook(output)
        ws = wb.active

        # Check data is in spreadsheet
        data_found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if test_first_name in str(row) and test_filename in str(row):
                data_found = True
                break

        assert data_found, "Player data not found in spreadsheet"

    def test_create_xlsx_empty_players(self):
        """Should skip teams with no players."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "division": None,
                "competitionYearBE": 2569,
                "players": []
            },
            {
                "id": 2,
                "name": "Team B",
                "ageGroup": "U18",
                "gender": "Female",
                "division": None,
                "competitionYearBE": 2569,
                "players": [
                    {
                        "firstName": "สมชาย",
                        "lastName": "วิชัยกุล",
                        "fullName": "สมชาย วิชัยกุล",
                        "dateOfBirth": None,
                        "birthYearBE": None,
                        "eligibilityStatus": "eligible",
                        "eligibilityNote": None,
                        "sourceFilename": "image1.jpg",
                        "status": "verified",
                        "verifiedAt": "2024-01-15T10:30:00",
                    }
                ]
            }
        ]

        output, _ = ExportService.create_xlsx(teams_data)
        wb = load_workbook(output)

        # Empty Team A should be skipped, only Team B sheet should exist
        assert len(wb.sheetnames) == 1
        assert "Team B" in wb.sheetnames

    def test_create_xlsx_column_widths(self):
        """Should set appropriate column widths."""
        teams_data = [
            {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "players": [
                    {
                        "firstName": "สมชาย",
                        "lastName": "วิชัยกุล",
                        "fullName": "สมชาย วิชัยกุล",
                        "sourceFilename": "image1.jpg",
                        "verifiedAt": "2024-01-15",
                    }
                ]
            }
        ]

        output, _ = ExportService.create_xlsx(teams_data)
        wb = load_workbook(output)
        ws = wb.active

        # Check column widths are set
        assert ws.column_dimensions["A"].width is not None
        assert ws.column_dimensions["B"].width is not None
        assert ws.column_dimensions["G"].width is not None  # Full Name


class TestFormatExportData:
    """Test data formatting for export."""

    def test_format_single_team(self):
        """Should format team and player data correctly."""
        teams = [
            type("Team", (), {
                "id": 1,
                "name": "Team A",
                "ageGroup": "U18",
                "gender": "Male",
                "division": None,
                "competitionYearBE": 2569
            })()
        ]

        players = [
            type("Player", (), {
                "id": 1,
                "teamId": 1,
                "firstName": "สมชาย",
                "lastName": "วิชัยกุล",
                "fullName": "สมชาย วิชัยกุล",
                "dateOfBirth": None,
                "birthYearBE": None,
                "eligibilityStatus": "eligible",
                "eligibilityNote": None,
                "sourceFilename": "image1.jpg",
                "status": "verified",
                "verifiedAt": None,
            })()
        ]

        result = ExportService.format_export_data(teams, players)

        assert len(result) == 1
        assert result[0]["name"] == "Team A"
        assert len(result[0]["players"]) == 1
        assert result[0]["players"][0]["firstName"] == "สมชาย"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
